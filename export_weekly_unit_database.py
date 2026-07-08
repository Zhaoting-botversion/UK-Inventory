from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
DB_PATH = SCRIPT_DIR / "inventory_units.sqlite"
REPORT_PATH = SCRIPT_DIR / "unit_change_import_report.json"
OUTPUT_DIR = SCRIPT_DIR / "outputs"
OUTPUT_PATH = OUTPUT_DIR / "weekly_unit_change_database.xlsx"


def clean_project_name(value: str) -> str:
    return re.sub(r"\s+·\s+.*$", "", value or "").strip()


def excel_price(value: object) -> int | None:
    if value in (None, ""):
        return None
    match = re.search(r"[\d,]+(?:\.\d+)?", str(value))
    if not match:
        return None
    try:
        return int(float(match.group(0).replace(",", "")))
    except ValueError:
        return None


def is_real_unit(value: object) -> bool:
    unit = str(value or "").strip()
    if not unit or not re.search(r"\d", unit):
        return False
    return unit.lower() not in {"plot", "plot no.", "plot no", "home no.", "home no", "unit", "unit no"}


def version_date(label: str, source_file: str) -> str:
    text = f"{label or ''} {source_file or ''}"
    patterns = [
        r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b",
        r"\b\d{1,2}(?:st|nd|rd|th)?\s+(?:Jan|Feb|Mar|Apr|May|Jun|June|Jul|July|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\s+\d{2,4}\b",
        r"\b(?:Jan|Feb|Mar|Apr|May|Jun|June|Jul|July|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\s+\d{2,4}\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return match.group(0)
    return label or source_file or ""


def fetch_rows() -> tuple[list[dict], list[dict], list[dict]]:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        events = [
            dict(row)
            for row in conn.execute(
                """
                SELECT
                    e.project_name,
                    e.unit,
                    e.floor,
                    e.bedroom,
                    e.aspect,
                    e.old_price,
                    old_v.version_label AS old_version_label,
                    old_v.source_file AS old_source_file,
                    e.new_price,
                    new_v.version_label AS new_version_label,
                    new_v.source_file AS new_source_file,
                    e.change_type,
                    e.old_status,
                    e.new_status,
                    e.price_change,
                    e.created_at
                FROM unit_change_events e
                LEFT JOIN pricelist_versions old_v ON old_v.id = e.old_version_id
                LEFT JOIN pricelist_versions new_v ON new_v.id = e.new_version_id
                ORDER BY e.project_name, e.change_type, e.unit
                """
            )
        ]
        current_units = [
            dict(row)
            for row in conn.execute(
                """
                WITH latest AS (
                    SELECT project_name, MAX(id) AS version_id
                    FROM pricelist_versions
                    GROUP BY project_name
                )
                SELECT
                    s.project_name,
                    s.unit,
                    s.floor,
                    s.bedroom,
                    s.aspect,
                    s.price,
                    s.status,
                    v.version_label,
                    v.source_file,
                    v.extracted_at
                FROM latest
                JOIN unit_snapshots s ON s.version_id = latest.version_id
                JOIN pricelist_versions v ON v.id = s.version_id
                ORDER BY s.project_name, s.unit
                """
            )
        ]
        versions = [
            dict(row)
            for row in conn.execute(
                """
                SELECT project_name, source_file, version_label, extracted_at, unit_count, parse_note
                FROM pricelist_versions
                ORDER BY extracted_at DESC, project_name, source_file
                """
            )
        ]
    return events, current_units, versions


def append_table(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        column = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column].width = min(max(max_len + 2, 10), 42)


EVENT_HEADERS = ["楼盘名", "房号", "楼层", "居室", "朝向", "原售价", "原售价生效日期", "现售价", "现售价生效日期", "变化类型", "原状态", "现状态", "价格变化", "入库时间"]


def event_row(row: dict) -> list[object]:
    return [
        clean_project_name(row["project_name"]),
        row["unit"],
        row["floor"],
        row["bedroom"],
        row["aspect"],
        excel_price(row["old_price"]),
        version_date(row["old_version_label"], row["old_source_file"]),
        excel_price(row["new_price"]),
        version_date(row["new_version_label"], row["new_source_file"]),
        row["change_type"],
        row["old_status"],
        row["new_status"],
        row["price_change"],
        row["created_at"],
    ]


def event_bucket(row: dict) -> str:
    change_type = row.get("change_type", "")
    status_text = f"{row.get('new_status', '')} {row.get('new_price', '')}".lower()
    if change_type == "PRICE_DROP":
        return "降价房源"
    if change_type in {"NEW_RELEASE", "BACK_ON_MARKET"}:
        return "新增房源"
    if change_type == "SOLD" or any(token in status_text for token in ["reserved", "under offer", "on hold", "hold", "reservation"]):
        return "售出房源"
    return "其他变化"


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    events, current_units, versions = fetch_rows()
    report = json.loads(REPORT_PATH.read_text(encoding="utf-8")) if REPORT_PATH.exists() else []
    events = [row for row in events if is_real_unit(row.get("unit"))]
    current_units = [row for row in current_units if is_real_unit(row.get("unit"))]

    wb = Workbook()
    ws = wb.active
    ws.title = "房源变化"
    append_table(
        ws,
        EVENT_HEADERS,
        [event_row(row) for row in events],
    )

    for sheet_name in ["降价房源", "新增房源", "售出房源", "其他变化"]:
        sheet = wb.create_sheet(sheet_name)
        append_table(
            sheet,
            EVENT_HEADERS,
            [event_row(row) for row in events if event_bucket(row) == sheet_name],
        )

    ws2 = wb.create_sheet("当前房源数据库")
    append_table(
        ws2,
        ["楼盘名", "房号", "楼层", "居室", "朝向", "现售价", "现状态", "现售价生效日期", "来源价单", "入库时间"],
        [
            [
                clean_project_name(row["project_name"]),
                row["unit"],
                row["floor"],
                row["bedroom"],
                row["aspect"],
                excel_price(row["price"]),
                row["status"],
                version_date(row["version_label"], row["source_file"]),
                row["source_file"],
                row["extracted_at"],
            ]
            for row in current_units
        ],
    )

    ws3 = wb.create_sheet("价单导入情况")
    append_table(
        ws3,
        ["楼盘名", "阶段/楼栋", "旧价单", "新价单", "旧房源数", "新房源数", "变化数", "旧价单错误", "新价单错误"],
        [
            [
                row.get("project", ""),
                row.get("project_key", ""),
                row.get("old_file", ""),
                row.get("new_file", ""),
                row.get("old_units", ""),
                row.get("new_units", ""),
                row.get("events", ""),
                row.get("old_error", ""),
                row.get("new_error", ""),
            ]
            for row in report
        ],
    )

    ws4 = wb.create_sheet("版本记录")
    append_table(
        ws4,
        ["楼盘名", "来源价单", "版本标签", "入库时间", "房源数", "解析备注"],
        [
            [
                clean_project_name(row["project_name"]),
                row["source_file"],
                row["version_label"],
                row["extracted_at"],
                row["unit_count"],
                row["parse_note"],
            ]
            for row in versions
        ],
    )

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, int) and ("售价" in str(sheet.cell(1, cell.column).value) or "价格变化" in str(sheet.cell(1, cell.column).value)):
                    cell.number_format = '£#,##0'
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)
    print(json.dumps({"events": len(events), "current_units": len(current_units), "versions": len(versions), "imports": len(report)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
